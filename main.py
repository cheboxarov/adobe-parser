import asyncio
from openpyxl import load_workbook
from adobe import AsyncClient
import os


async def procces_file(file_name: str):
    print(f"Обрабатываю файл {file_name}")
    wb = load_workbook(file_name)
    ws = wb.active

    tasks = []

    empty_cells = 0

    for row in ws.iter_rows():
        url = row[0].value
        name = row[1].value

        if name is None:
            empty_cells += 1
            tasks.append(fill_asset_name(url, row))

    if empty_cells == 0:
        print("Нет пустых ячеек имен")
        return
    print(f"Обрабатываю {empty_cells} страниц")
    await asyncio.gather(*tasks)

    wb.save('test.xlsx')
    print(f"Изменения сохранены в {file_name}")


async def main():
    current_dir = os.getcwd()
    xlsx_files = [f for f in os.listdir(current_dir) if f.endswith('.xlsx')]

    if not xlsx_files:
        print("Не найдены таблицы")

    for file in xlsx_files:
        await procces_file(file)


async def fill_asset_name(url, row):
    try:
        adobe_client = AsyncClient()
        asset_name = await adobe_client.get_asset_name(url)
        row[1].value = asset_name
    except:
        row[1].value = "Error!"


if __name__ == "__main__":
    asyncio.get_event_loop().run_until_complete(main())
